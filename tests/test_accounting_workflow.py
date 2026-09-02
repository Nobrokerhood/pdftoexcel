import io
from dataclasses import replace

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.app_factory import create_app
from app.auth.google_auth import VerifiedGoogleUser
from app.core.config import Settings
from app.google.sheet_schemas import schema_for
from app.processing.stores import GoogleSheetsProcessingJobStore


def settings(max_retries=2) -> Settings:
    return Settings(
        google_client_id="client-id",
        frontend_google_client_id="client-id",
        allowed_email_domain="nobroker.in",
        allow_domain_wide_access=False,
        session_inactivity_seconds=1200,
        session_heartbeat_grace_seconds=120,
        ai_verification_max_retries=max_retries,
        allow_dev_login=False,
        gemini_api_key=None,
        gemini_model="gemini-2.5-flash",
        max_file_size_mb=10,
        google_service_account_json=None,
        google_service_account_file=None,
        google_accounting_spreadsheet_id=None,
        google_user_master_sheet_id="users",
        google_login_audit_sheet_id="login",
        google_api_usage_sheet_id="usage",
        google_session_log_sheet_id="sessions",
        google_activity_log_sheet_id="activity",
        google_processing_log_sheet_id="processing",
        google_template_master_sheet_id="templates",
        google_folder_config_sheet_id="folders",
        google_mapping_master_sheet_id="mapping",
        google_drive_root_folder_id="root",
        google_login_audit_sheet_name="Accounting_AI_Login_Audit",
        google_api_usage_sheet_name="API_Usage_Report",
        cors_allowed_origins=("http://localhost:5000",),
    )


class FakeSheetsService:
    def __init__(self, records):
        self.records = records
        self.appended = []
        self.updated = []

    def read_records(self, spreadsheet_id, worksheet_name="Sheet1", use_cache=True):
        return [row.copy() for row in self.records.get(spreadsheet_id, [])]

    def read_table(self, table_key, use_cache=True):
        table_map = {
            "user_master": "users",
            "folder_config": "folders",
            "template_master": "templates",
            "mapping_master": "mapping",
            "login_audit": "login",
            "activity_log": "activity",
            "session_log": "sessions",
            "processing_log": "processing",
            "job_state": "job_state",
        }
        return self.read_records(table_map.get(table_key, table_key))

    def append_row(self, spreadsheet_id, values, worksheet_name="Sheet1"):
        self.appended.append((spreadsheet_id, values))
        return True

    def append_table_row(self, table_key, values):
        table_map = {
            "login_audit": "login",
            "activity_log": "activity",
            "session_log": "sessions",
            "processing_log": "processing",
            "job_state": "job_state",
        }
        return self.append_row(table_map.get(table_key, table_key), values)

    def lookup_row_by_key(self, spreadsheet_id, key_column, key_value, worksheet_name="Sheet1"):
        for record in self.read_records(spreadsheet_id, worksheet_name):
            if str(record.get(key_column, "")).strip().lower() == key_value.lower():
                return record
        return None

    def lookup_table_row_by_key(self, table_key, key_column, key_value):
        for record in self.read_table(table_key):
            if str(record.get(key_column, "")).strip().lower() == key_value.lower():
                return record
        return None

    def update_row_by_key(self, spreadsheet_id, key_column, key_value, updates, worksheet_name="Sheet1"):
        self.updated.append((spreadsheet_id, key_column, key_value, updates))
        return True

    def update_table_row_by_key(self, table_key, key_column, key_value, updates):
        table_map = {
            "login_audit": "login",
            "processing_log": "processing",
            "job_state": "job_state",
        }
        return self.update_row_by_key(table_map.get(table_key, table_key), key_column, key_value, updates)


class FakeVerifier:
    def verify(self, credential):
        return VerifiedGoogleUser("active@nobroker.in", "Active User")


class FakeDriveService:
    def __init__(self):
        self.uploads = []
        self.moves = []
        self.downloads = {}

    def upload_file(self, filename, content, folder_id, mime_type="application/octet-stream"):
        file_id = f"{folder_id}-{len(self.uploads) + 1}"
        if isinstance(content, bytes):
            stored_content = content
        else:
            stored_content = content.read()
        self.uploads.append(
            {
                "filename": filename,
                "content": stored_content,
                "folder_id": folder_id,
                "mime_type": mime_type,
                "file_id": file_id,
            }
        )
        self.downloads[file_id] = stored_content
        return file_id

    def move_file(self, file_id, folder_id):
        self.moves.append((file_id, folder_id))
        return True

    def download_file(self, file_id):
        return self.downloads[file_id]


class StaticExtractionProvider:
    def __init__(self, data):
        self.data = data

    def extract(self, source_bytes, purpose, template):
        return self.data[purpose].copy()


class StaticVerificationProvider:
    def __init__(self, results=None):
        self.results = list(results or [{"overall_status": "PASSED", "fields": []}])
        self.calls = 0

    def verify(self, source_bytes, purpose, template, extracted_data):
        index = min(self.calls, len(self.results) - 1)
        self.calls += 1
        return self.results[index]


class StaticRepairProvider:
    def __init__(self, repaired):
        self.repaired = repaired
        self.calls = 0

    def repair(self, source_bytes, purpose, template, extracted_data, verification_result):
        self.calls += 1
        return self.repaired.copy()


def records(include_mapping=True):
    mapping = []
    if include_mapping:
        mapping = [
            {"Purpose": "MEMBER_RECEIPT", "Type": "BANK", "Source Value": "HDFC", "Target Value": "HDFC001", "Active": "true"},
            {"Purpose": "MEMBER_RECEIPT", "Type": "BILL_HEAD", "Source Value": "Maintenance", "Target Value": "MAINT", "Active": "true"},
            {"Purpose": "VENDOR_INVOICE", "Type": "VENDOR", "Source Value": "ABC Plumbing Services Pvt Ltd", "Target Value": "VEND-ABC", "Active": "true"},
            {"Purpose": "VENDOR_INVOICE", "Type": "EXPENSE", "Source Value": "Plumbing work", "Target Value": "REPAIR", "Active": "true"},
        ]
    return {
        "users": [
            {"Email": "active@nobroker.in", "Name": "Active User", "Role": "ADMIN", "Active": "true"}
        ],
        "folders": [
            {"Purpose": "MEMBER_RECEIPT", "Incoming Folder ID": "member-in", "Review Folder ID": "member-review", "Completed Folder ID": "member-done", "Output Folder ID": "member-out", "Active": "true"},
            {"Purpose": "VENDOR_INVOICE", "Incoming Folder ID": "vendor-in", "Review Folder ID": "vendor-review", "Completed Folder ID": "vendor-done", "Output Folder ID": "vendor-out", "Active": "true"},
        ],
        "mapping": mapping,
    }


MEMBER_DATA = {
    "payment_type": "UPI",
    "bank_name_or_code": "HDFC",
    "reference_number": "UPI123456",
    "tower": "A",
    "flat": "101",
    "bill_head": "Maintenance",
    "amount": "5000",
    "transaction_date": "25-Aug-2026",
    "comments": "SYNTHETIC TEST DATA",
    "meter_number": None,
    "cheque_issuer_bank": None,
    "cheque_date": None,
}

VENDOR_DATA = {
    "bill_number": "INV-418",
    "bill_date": "25-Aug-2026",
    "vendor_code": None,
    "vendor_name": "ABC Plumbing Services Pvt Ltd",
    "due_date": None,
    "narration": "SYNTHETIC TEST DATA",
    "cgst_amount": "900",
    "sgst_amount": "900",
    "igst_amount": "0",
    "tds_amount": "0",
    "expenses": [
        {"expense_code": None, "expense_description": "Plumbing work", "expense_amount": "10000"}
    ],
}


def client_for(data=None, sheet_records=None, verify_results=None, repaired=None, max_retries=2):
    drive = FakeDriveService()
    sheets = FakeSheetsService(sheet_records or records())
    app = create_app(
        settings=settings(max_retries),
        sheets_service=sheets,
        drive_service=drive,
        google_token_verifier=FakeVerifier(),
        extraction_provider=StaticExtractionProvider(data or {"MEMBER_RECEIPT": MEMBER_DATA, "VENDOR_INVOICE": VENDOR_DATA}),
        verification_provider=StaticVerificationProvider(verify_results),
        repair_provider=StaticRepairProvider(repaired or MEMBER_DATA),
    )
    return TestClient(app), app, drive, sheets


class PersistentStateSheets(FakeSheetsService):
    def __init__(self, records):
        super().__init__(records)
        self.records.setdefault("job_state", [])

    def append_table_row(self, table_key, values):
        if table_key == "job_state":
            headers = schema_for("job_state").headers
            self.records["job_state"].append(dict(zip(headers, values)))
            self.appended.append(("job_state", values))
            return True
        return super().append_table_row(table_key, values)

    def update_table_row_by_key(self, table_key, key_column, key_value, updates):
        if table_key == "job_state":
            for record in self.records["job_state"]:
                if str(record.get(key_column, "")).strip().lower() == key_value.lower():
                    record.update(updates)
                    self.updated.append(("job_state", key_column, key_value, updates))
                    return True
            return False
        return super().update_table_row_by_key(table_key, key_column, key_value, updates)


def persistent_client_for(sheet_records, drive=None, job_store=None):
    drive = drive or FakeDriveService()
    sheets = PersistentStateSheets(sheet_records)
    app_settings = replace(settings(), google_accounting_spreadsheet_id="shared")
    app = create_app(
        settings=app_settings,
        sheets_service=sheets,
        drive_service=drive,
        google_token_verifier=FakeVerifier(),
        extraction_provider=StaticExtractionProvider(
            {"MEMBER_RECEIPT": MEMBER_DATA, "VENDOR_INVOICE": VENDOR_DATA}
        ),
        verification_provider=StaticVerificationProvider(),
        repair_provider=StaticRepairProvider(MEMBER_DATA),
        job_store=job_store or GoogleSheetsProcessingJobStore(sheets),
    )
    return TestClient(app), app, drive, sheets


def token(client):
    response = client.post("/auth/google-login", json={"credential": "token"})
    assert response.status_code == 200
    return response.json()["session_token"]


def headers(session_token):
    return {"Authorization": f"Bearer {session_token}"}


def start_job(client, session_token, purpose="MEMBER_RECEIPT"):
    return client.post(
        "/processing/jobs",
        headers=headers(session_token),
        data={"purpose": purpose},
        files={"file": ("synthetic.pdf", b"SYNTHETIC TEST DATA", "application/pdf")},
    )


def test_member_receipt_happy_path_generates_exact_xlsx_after_approval():
    client, app, drive, _ = client_for()
    session_token = token(client)

    started = start_job(client, session_token)

    assert started.status_code == 200
    body = started.json()
    assert body["overall_status"] == "NEEDS_REVIEW"
    assert body["verification_status"] == "PASSED"
    assert body["mapping_status"] == "MAPPED"
    assert body["validation_status"] == "PASSED"
    assert drive.uploads[0]["folder_id"] == "member-in"

    approved = client.post(f"/processing/jobs/{body['job_id']}/approve", headers=headers(session_token))
    assert approved.status_code == 200
    approved_body = approved.json()
    assert approved_body["overall_status"] == "COMPLETED"
    assert drive.uploads[1]["folder_id"] == "member-out"
    assert drive.moves == [(body["source_drive_file_id"], "member-done")]

    downloaded = client.get(f"/processing/jobs/{body['job_id']}/download", headers=headers(session_token))
    wb = load_workbook(io.BytesIO(downloaded.content))
    columns = [cell.value for cell in next(wb.active.iter_rows(max_row=1))]
    assert columns == [
        "Payment Type",
        "Society Bank Name/Bank code",
        "Cheque/Ref No",
        "Tower No",
        "Flat No",
        "Bill Head",
        "Amount",
        "Transaction Date",
        "Comments",
        "Meter No",
        "Cheque Issuer Bank",
        "Cheque Date",
    ]


def test_vendor_invoice_happy_path_generates_exact_xlsx_after_approval():
    client, _, drive, _ = client_for()
    session_token = token(client)

    started = start_job(client, session_token, "VENDOR_INVOICE")
    body = started.json()

    assert started.status_code == 200
    assert body["mapping_status"] == "MAPPED"
    assert body["validation_status"] == "PASSED"
    assert drive.uploads[0]["folder_id"] == "vendor-in"

    approved = client.post(f"/processing/jobs/{body['job_id']}/approve", headers=headers(session_token))
    assert approved.status_code == 200
    downloaded = client.get(f"/processing/jobs/{body['job_id']}/download", headers=headers(session_token))
    wb = load_workbook(io.BytesIO(downloaded.content))
    columns = [cell.value for cell in next(wb.active.iter_rows(max_row=1))]
    assert columns == [
        "Bill Number",
        "Bill Date",
        "Vendor Code",
        "Due Date",
        "Narration",
        "CGST Amount",
        "SGST Amount",
        "IGST Amount",
        "TDS Amount",
        "Expense Code",
        "Expense Amount",
    ]
    assert drive.uploads[1]["folder_id"] == "vendor-out"


def test_verification_mismatch_repairs_then_passes():
    bad = MEMBER_DATA.copy()
    bad["amount"] = "18800"
    repaired = MEMBER_DATA.copy()
    repaired["amount"] = "11800"
    client, _, _, _ = client_for(
        data={"MEMBER_RECEIPT": bad, "VENDOR_INVOICE": VENDOR_DATA},
        repaired=repaired,
        verify_results=[
            {
                "overall_status": "FAILED",
                "fields": [
                    {
                        "field": "amount",
                        "extracted_value": "18800",
                        "verified_value": "11800",
                        "status": "MISMATCH",
                        "confidence": 0.98,
                        "evidence": "Total = 11800",
                    }
                ],
            },
            {"overall_status": "PASSED", "fields": []},
        ],
    )
    session_token = token(client)

    response = start_job(client, session_token)

    assert response.status_code == 200
    body = response.json()
    assert body["extraction_attempt"] == 2
    assert body["verification_status"] == "PASSED"
    assert body["extracted_data"]["amount"] == "11800"


def test_verification_retry_exhausted_needs_review():
    client, _, _, _ = client_for(
        verify_results=[{"overall_status": "FAILED", "fields": [{"field": "amount", "status": "MISMATCH"}]}],
        max_retries=0,
    )
    session_token = token(client)

    response = start_job(client, session_token)

    assert response.status_code == 200
    body = response.json()
    assert body["overall_status"] == "NEEDS_REVIEW"
    assert body["verification_status"] == "FAILED"


def test_missing_mapping_needs_review_then_resolution_resumes_validation():
    client, _, _, _ = client_for(sheet_records=records(include_mapping=False))
    session_token = token(client)
    started = start_job(client, session_token, "VENDOR_INVOICE")
    body = started.json()

    assert body["mapping_status"] == "NEEDS_MAPPING"
    assert body["overall_status"] == "NEEDS_REVIEW"

    resolved = client.post(
        f"/processing/jobs/{body['job_id']}/mapping",
        headers=headers(session_token),
        json={
            "resolutions": [
                {"type": "VENDOR", "source_value": "ABC Plumbing Services Pvt Ltd", "target_value": "VEND-ABC"},
                {"type": "EXPENSE", "source_value": "Plumbing work", "target_value": "REPAIR"},
            ]
        },
    )

    assert resolved.status_code == 200
    resolved_body = resolved.json()
    assert resolved_body["mapping_status"] == "MAPPED"
    assert resolved_body["validation_status"] == "PASSED"


def test_human_edit_reruns_validation():
    invalid = MEMBER_DATA.copy()
    invalid["amount"] = None
    client, _, _, _ = client_for(data={"MEMBER_RECEIPT": invalid, "VENDOR_INVOICE": VENDOR_DATA})
    session_token = token(client)
    started = start_job(client, session_token)
    body = started.json()
    assert body["validation_status"] == "BLOCKED"

    edited = client.post(
        f"/processing/jobs/{body['job_id']}/corrections",
        headers=headers(session_token),
        json={"corrections": {"amount": "5000"}},
    )

    assert edited.status_code == 200
    assert edited.json()["validation_status"] == "PASSED"
    assert edited.json()["human_corrections"][0]["field"] == "amount"


def test_human_reject_generates_no_output():
    client, _, _, _ = client_for()
    session_token = token(client)
    started = start_job(client, session_token)
    body = started.json()

    rejected = client.post(f"/processing/jobs/{body['job_id']}/reject", headers=headers(session_token))
    downloaded = client.get(f"/processing/jobs/{body['job_id']}/download", headers=headers(session_token))

    assert rejected.status_code == 200
    assert rejected.json()["overall_status"] == "REJECTED"
    assert downloaded.status_code == 404


def test_audit_sequence_contains_expected_actions():
    client, _, _, sheets = client_for()
    session_token = token(client)
    started = start_job(client, session_token)
    client.post(f"/processing/jobs/{started.json()['job_id']}/approve", headers=headers(session_token))

    actions = [row[1][4] for row in sheets.appended if row[0] == "activity"]
    assert "LOGIN" in actions
    assert "PURPOSE_SELECTED" in actions
    assert "FILE_UPLOAD" in actions
    assert "EXTRACTION_STARTED" in actions
    assert "EXTRACTION_COMPLETED" in actions
    assert "AI_VERIFICATION_STARTED" in actions
    assert "AI_VERIFICATION_PASSED" in actions
    assert "HUMAN_APPROVED" in actions
    assert "EXCEL_GENERATED" in actions


def test_job_state_restores_after_restart_approval_and_download_use_drive_id():
    sheet_records = records()
    client1, _, drive1, sheets1 = persistent_client_for(sheet_records)
    token1 = token(client1)

    started = start_job(client1, token1)
    body = started.json()
    job_id = body["job_id"]

    assert started.status_code == 200
    assert body["current_step"] == "HUMAN_REVIEW"
    assert sheet_records["job_state"][0]["Current Step"] == "HUMAN_REVIEW"
    assert sheet_records["job_state"][0]["Overall Status"] == "NEEDS_REVIEW"

    client2, _, drive2, _ = persistent_client_for(sheet_records)
    token2 = token(client2)

    listed = client2.get("/processing/jobs", headers=headers(token2))
    restored = client2.get(f"/processing/jobs/{job_id}", headers=headers(token2))

    assert listed.status_code == 200
    assert [job["job_id"] for job in listed.json()["jobs"]] == [job_id]
    assert restored.status_code == 200
    assert restored.json()["output_folder_id"] == "member-out"

    approved = client2.post(f"/processing/jobs/{job_id}/approve", headers=headers(token2))
    assert approved.status_code == 200
    approved_body = approved.json()
    output_file_id = approved_body["output_drive_file_id"]
    assert approved_body["overall_status"] == "COMPLETED"
    assert output_file_id
    assert drive2.moves == [(body["source_drive_file_id"], "member-done")]

    drive3 = FakeDriveService()
    drive3.downloads[output_file_id] = drive2.downloads[output_file_id]
    client3, _, _, _ = persistent_client_for(sheet_records, drive=drive3)
    token3 = token(client3)

    downloaded = client3.get(f"/processing/jobs/{job_id}/download", headers=headers(token3))
    wb = load_workbook(io.BytesIO(downloaded.content))

    assert downloaded.status_code == 200
    assert drive1.uploads[0]["folder_id"] == "member-in"
    assert wb.active["A1"].value == "Payment Type"


def test_google_sheets_job_store_create_update_list_and_load_state():
    sheet_records = records()
    sheets = PersistentStateSheets(sheet_records)
    store = GoogleSheetsProcessingJobStore(sheets)

    job = store.create_job(
        "session-1",
        "active@nobroker.in",
        "MEMBER_RECEIPT",
        "NBH_MEMBER_RECEIPT_V1",
        "receipt.pdf",
        "application/pdf",
        b"source bytes",
    )
    job.current_step = "HUMAN_REVIEW"
    job.overall_status = "NEEDS_REVIEW"
    store.update_job(job)

    loaded = store.load_state(job.job_id)
    restored = GoogleSheetsProcessingJobStore(sheets).get(job.job_id)

    assert loaded["current_step"] == "HUMAN_REVIEW"
    assert restored.source_bytes == b""
    assert restored.current_step == "HUMAN_REVIEW"
    assert [item.job_id for item in store.list_jobs("active@nobroker.in")] == [job.job_id]
