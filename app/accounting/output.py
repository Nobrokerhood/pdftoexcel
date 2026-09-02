import io
from decimal import Decimal

from openpyxl import Workbook

from app.accounting.purposes import MEMBER_RECEIPT, VENDOR_INVOICE
from app.accounting.templates import TemplateDefinition


class OutputGenerationError(RuntimeError):
    pass


def _value(value):
    if isinstance(value, Decimal):
        return float(value)
    return value if value is not None else ""


class TemplateOutputGenerator:
    def generate_xlsx(
        self,
        purpose: str,
        template: TemplateDefinition,
        data: dict,
        job_id: str,
    ) -> tuple[str, bytes]:
        purpose = purpose.upper()
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = template.template_code[:31]
        sheet.append(list(template.fields))

        if purpose == MEMBER_RECEIPT:
            sheet.append(
                [
                    _value(data.get("payment_type")),
                    _value(data.get("bank_name_or_code")),
                    _value(data.get("reference_number")),
                    _value(data.get("tower")),
                    _value(data.get("flat")),
                    _value(data.get("bill_head")),
                    _value(data.get("amount")),
                    _value(data.get("transaction_date")),
                    _value(data.get("comments")),
                    _value(data.get("meter_number")),
                    _value(data.get("cheque_issuer_bank")),
                    _value(data.get("cheque_date")),
                ]
            )
        elif purpose == VENDOR_INVOICE:
            expenses = data.get("expenses") or [{}]
            for expense in expenses:
                sheet.append(
                    [
                        _value(data.get("bill_number")),
                        _value(data.get("bill_date")),
                        _value(data.get("vendor_code")),
                        _value(data.get("due_date")),
                        _value(data.get("narration")),
                        _value(data.get("cgst_amount")),
                        _value(data.get("sgst_amount")),
                        _value(data.get("igst_amount")),
                        _value(data.get("tds_amount")),
                        _value(expense.get("expense_code")),
                        _value(expense.get("expense_amount")),
                    ]
                )
        else:
            raise OutputGenerationError("Unsupported purpose.")

        for row in sheet.iter_rows():
            for cell in row:
                if "date" in str(sheet.cell(row=1, column=cell.column).value).lower():
                    cell.number_format = "dd-mmm-yyyy"
                if "amount" in str(sheet.cell(row=1, column=cell.column).value).lower():
                    cell.number_format = "#,##0.00"

        stream = io.BytesIO()
        workbook.save(stream)
        filename = f"{template.template_code}_{job_id}.xlsx"
        return filename, stream.getvalue()
